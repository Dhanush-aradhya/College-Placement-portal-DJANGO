# home/views.py
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout # Not using these for DUMMY login
from django.contrib.auth.decorators import login_required # Will use a custom decorator or check
from django.contrib import messages
from django.contrib.auth import logout
from django.core.exceptions import PermissionDenied
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash 
from .models import StudentProfile,Domain, Language # Import your StudentProfile model
from .forms import StudentProfileForm # Import your new form
from django.db.models import Q
from django.http import HttpResponse
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment # For basic styling like bold headers
from openpyxl.utils import get_column_letter # To convert column index to letter
from io import BytesIO

# Custom decorator to check our dummy session login



def index(request):
    return render(request, 'index.html')

@login_required(login_url='userlogin')
def userportal(request):
    if request.user.is_staff:
        messages.warning(request, "Staff members should use the admin portal.")
        return redirect('adminportal')

    student_profile = None
    try:
        # Fetch the student profile linked to the currently logged-in user
        student_profile = StudentProfile.objects.get(user=request.user)
    except StudentProfile.DoesNotExist:
        # No profile exists yet for this user.
        # You might want to add a message encouraging them to create one.
        messages.info(request, "Welcome! Please upload your details to complete your profile.")
        # No critical error, just means they haven't uploaded details yet.
        # The template will handle the 'student_profile is None' case.

    context = {
        'student_profile': student_profile,
        # 'user' is already available in the template context as request.user
    }
    return render(request, 'userportal.html', context)

# home/views.py
# ... (other imports remain the same) ...

@login_required(login_url='adminlogin')
def adminportal(request):
    if not request.user.is_staff:
        messages.error(request, "You do not have permission to access this portal.")
        return redirect('home')

    student_profiles_list = StudentProfile.objects.select_related('user').prefetch_related('domains', 'languages_known').order_by('full_name')

    # --- Get filter parameters ---
    department_filters = request.GET.getlist('departmentFilter')
    year_filters = request.GET.getlist('yearFilter')
    cgpa_filter_str = request.GET.get('cgpaFilter', '').strip()
    backlogs_filter_str = request.GET.get('backlogsFilter', '').strip()
    domain_filter_ids_str = request.GET.getlist('domainFilter')
    language_filter_ids_str = request.GET.getlist('languageFilter') # NEW: Language filter

    tenth_percentage_min_str = request.GET.get('tenthPercentageMin', '').strip()
    # REMOVED: tenth_percentage_max_str
    pre_uni_percentage_min_str = request.GET.get('preUniPercentageMin', '').strip()
    # REMOVED: pre_uni_percentage_max_str

    # --- Apply filters ---
    if department_filters:
        student_profiles_list = student_profiles_list.filter(department__in=department_filters)
    if year_filters:
        student_profiles_list = student_profiles_list.filter(current_year__in=year_filters)
    if cgpa_filter_str:
        try:
            cgpa_filter_val = float(cgpa_filter_str)
            student_profiles_list = student_profiles_list.filter(cgpa__gte=cgpa_filter_val)
        except ValueError:
            messages.error(request, f"Invalid Min CGPA value '{cgpa_filter_str}' ignored.")
    if backlogs_filter_str:
        try:
            backlogs_filter_val = int(backlogs_filter_str)
            student_profiles_list = student_profiles_list.filter(backlogs__lte=backlogs_filter_val)
        except ValueError:
            messages.error(request, f"Invalid Max Backlogs value '{backlogs_filter_str}' ignored.")

    domain_filter_ids_int = []
    if domain_filter_ids_str:
        for id_str in domain_filter_ids_str:
            if id_str.isdigit():
                domain_filter_ids_int.append(int(id_str))
        if domain_filter_ids_int:
            student_profiles_list = student_profiles_list.filter(domains__id__in=domain_filter_ids_int).distinct()
    
    # NEW: Apply Language Filter
    language_filter_ids_int = []
    if language_filter_ids_str:
        for id_str in language_filter_ids_str:
            if id_str.isdigit():
                language_filter_ids_int.append(int(id_str))
        if language_filter_ids_int:
            student_profiles_list = student_profiles_list.filter(languages_known__id__in=language_filter_ids_int).distinct()


    if tenth_percentage_min_str:
        try:
            val = float(tenth_percentage_min_str)
            student_profiles_list = student_profiles_list.filter(tenth_percentage__gte=val)
        except ValueError:
            messages.error(request, f"Invalid Min 10th % value '{tenth_percentage_min_str}' ignored.")
    # REMOVED: Max 10th % filter logic
    
    if pre_uni_percentage_min_str:
        try:
            val = float(pre_uni_percentage_min_str)
            student_profiles_list = student_profiles_list.filter(pre_university_percentage__gte=val)
        except ValueError:
            messages.error(request, f"Invalid Min 12th/Dip % value '{pre_uni_percentage_min_str}' ignored.")
    # REMOVED: Max 12th/Dip % filter logic

    current_filters = {
        'departmentFilter_list': department_filters,
        'yearFilter_list': year_filters,
        'cgpaFilter': cgpa_filter_str,
        'backlogsFilter': backlogs_filter_str,
        'domainFilter_ids': domain_filter_ids_int,
        'languageFilter_ids': language_filter_ids_int, # NEW
        'tenthPercentageMin': tenth_percentage_min_str,
        # REMOVED: 'tenthPercentageMax'
        'preUniPercentageMin': pre_uni_percentage_min_str,
        # REMOVED: 'preUniPercentageMax'
    }
    
    all_domains_for_filter = Domain.objects.all()
    all_languages_for_filter = Language.objects.all() # NEW: Get all languages

    context = {
        'student_profiles': student_profiles_list,
        'current_filters': current_filters,
        'all_domains': all_domains_for_filter,
        'all_languages': all_languages_for_filter, # NEW
        'department_choices': StudentProfile._meta.get_field('department').choices,
        'year_choices': StudentProfile._meta.get_field('current_year').choices,
    }
    return render(request, 'adminportal.html', context)

@login_required(login_url='userlogin')
def update_password(request):
    if request.user.is_staff:
        messages.info(request, "Staff password changes should be done via the admin portal or a dedicated staff password change page.")
        return redirect('adminportal') # Or wherever appropriate for staff

    if request.method == 'POST':
        # Pass the current user to the form!
        form = PasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save() # This saves the new password and hashes it
            # Important: Update the session hash to prevent the user from being logged out
            update_session_auth_hash(request, user)
            messages.success(request, 'Your password was successfully updated!')
            return redirect('userportal') # Or to a dedicated 'password_change_done' page
        else:
            # Form is not valid, errors will be displayed with the form
            messages.error(request, 'Please correct the error(s) below.')
    else:
        # Pass the current user to the form for GET requests as well
        form = PasswordChangeForm(user=request.user)

    return render(request, 'update-password.html', {'form': form})

@login_required(login_url='userlogin')
def upload_details(request):
    if request.user.is_staff: # Staff should not be uploading details like students
        messages.info(request, "This page is for student detail uploads.")
        return redirect('adminportal')

    try:
        # Try to get an existing profile for the logged-in user
        student_profile = StudentProfile.objects.get(user=request.user)
    except StudentProfile.DoesNotExist:
        # If no profile exists, set to None. A new one will be created on POST.
        student_profile = None

    if request.method == 'POST':
        # If updating an existing profile, pass the instance.
        # If creating a new one, student_profile will be None.
        form = StudentProfileForm(request.POST, request.FILES, instance=student_profile)
        if form.is_valid():
            # If it's a new profile, we need to associate it with the user
            # and potentially set the USN from the username before saving.
            profile = form.save(commit=False) # Don't save to DB yet
            profile.user = request.user      # Link the profile to the current user

            # Handle USN:
            # If USN is a field in StudentProfileForm and should be based on username
            # AND it's a new profile (no instance.pk yet)
            if not profile.pk: # This means it's a new profile
                # Assuming the username IS the USN. If usn field is not in form, this is fine.
                # If usn field IS in form and should be auto-set for new profiles:
                profile.usn = request.user.username # Set USN from username

            profile.save() # Now save the profile to the database
            form.save_m2m() # If you had ManyToManyFields, you'd call this after profile.save()

            messages.success(request, 'Your profile has been successfully updated!')
            return redirect('userportal') # Redirect to user portal after successful save
        else:
            # Form is not valid, errors will be displayed with the form by the template
            messages.error(request, 'Please correct the errors below.')
    else: # GET request
        # If profile exists, pre-fill the form with its data.
        # If not, an empty form will be shown.
        initial_data = {}
        if not student_profile: # If new profile, pre-fill USN from username
            initial_data['usn'] = request.user.username

        form = StudentProfileForm(instance=student_profile, initial=initial_data if not student_profile else None)

    # For the USN field, if the profile exists, make it readonly in the template
    # We can pass a flag to the template or handle it directly in the template
    usn_is_readonly = student_profile is not None

    return render(request, 'upload-details.html', {
        'form': form,
        'usn_is_readonly': usn_is_readonly,
        'current_usn': student_profile.usn if student_profile else request.user.username
    })

def userlogin_view(request):
    if request.user.is_authenticated: # Check if user is already logged in
        # If they are staff, maybe they landed here by mistake, send to admin portal
        if request.user.is_staff:
            return redirect('adminportal')
        return redirect('userportal')

    if request.method == "POST":
        uname = request.POST.get('username')
        pword = request.POST.get('password')

        user = authenticate(request, username=uname, password=pword)

        if user is not None:
            # Check if the user is staff. Regular users should not log in here if is_staff is True
            # (unless you allow staff to also have a "user" view, which is less common for distinct portals)
            if user.is_staff:
                messages.error(request, "Staff accounts should use the Admin Login.")
                return render(request, 'userlogin.html', {'error': "Please use Admin Login."})
            else:
                login(request, user)
                messages.success(request, f"Welcome back, {user.username}!")
                return redirect('userportal')
        else:
            messages.error(request, "Invalid username or password.")
            # No need to pass 'error' in context, messages framework handles it
            return render(request, 'userlogin.html')
    return render(request, 'userlogin.html')

def adminlogin_view(request):
    if request.user.is_authenticated and request.user.is_staff:
        return redirect('adminportal')
    elif request.user.is_authenticated and not request.user.is_staff:
        # A regular user is logged in and trying to access admin login.
        # Log them out and redirect to admin login or show an error.
        logout(request) # Log out the current non-staff user
        messages.info(request, "You have been logged out. Please log in with an admin account.")
        # Or: messages.error(request, "Only admin/staff can log in here.")
        # return redirect('home')

    if request.method == "POST":
        uname = request.POST.get('username')
        pword = request.POST.get('password')

        user = authenticate(request, username=uname, password=pword)

        if user is not None:
            if user.is_staff: # Only allow staff members
                login(request, user)
                messages.success(request, f"Admin login successful, {user.username}!")
                return redirect('adminportal')
            else:
                messages.error(request, "This account does not have admin/staff privileges.")
                return render(request, 'adminlogin.html')
        else:
            messages.error(request, "Invalid admin username or password.")
            return render(request, 'adminlogin.html')
    return render(request, 'adminlogin.html')

# from django.contrib.auth import logout # Already imported at the top

def logout_view(request):
    current_user_username = request.user.username # Get username before logout for message
    logout(request) # This clears the authentication session data
    messages.success(request, f"User '{current_user_username}' has been successfully logged out.")
    return redirect('home')

from django.http import HttpResponse # Add this import
# ... other imports ...

@login_required(login_url='adminlogin')
def download_students_excel_view(request):
    if not request.user.is_staff:
        messages.error(request, "You do not have permission to perform this action.")
        return redirect('adminportal')

    # --- Start: Filtering logic (copied and adapted from adminportal view) ---
    student_profiles_list = StudentProfile.objects.select_related('user').prefetch_related('domains', 'languages_known').order_by('full_name')

    department_filters = request.GET.getlist('departmentFilter')
    year_filters = request.GET.getlist('yearFilter')
    cgpa_filter_str = request.GET.get('cgpaFilter', '').strip()
    backlogs_filter_str = request.GET.get('backlogsFilter', '').strip()
    domain_filter_ids_str = request.GET.getlist('domainFilter')
    language_filter_ids_str = request.GET.getlist('languageFilter')
    tenth_percentage_min_str = request.GET.get('tenthPercentageMin', '').strip()
    pre_uni_percentage_min_str = request.GET.get('preUniPercentageMin', '').strip()

    if department_filters:
        student_profiles_list = student_profiles_list.filter(department__in=department_filters)
    if year_filters:
        student_profiles_list = student_profiles_list.filter(current_year__in=year_filters)
    if cgpa_filter_str:
        try:
            cgpa_filter_val = float(cgpa_filter_str)
            student_profiles_list = student_profiles_list.filter(cgpa__gte=cgpa_filter_val)
        except ValueError: pass # Silently ignore invalid filter
    if backlogs_filter_str:
        try:
            backlogs_filter_val = int(backlogs_filter_str)
            student_profiles_list = student_profiles_list.filter(backlogs__lte=backlogs_filter_val)
        except ValueError: pass
    
    domain_filter_ids_int = [int(id_str) for id_str in domain_filter_ids_str if id_str.isdigit()]
    if domain_filter_ids_int:
        student_profiles_list = student_profiles_list.filter(domains__id__in=domain_filter_ids_int).distinct()
    
    language_filter_ids_int = [int(id_str) for id_str in language_filter_ids_str if id_str.isdigit()]
    if language_filter_ids_int:
        student_profiles_list = student_profiles_list.filter(languages_known__id__in=language_filter_ids_int).distinct()

    if tenth_percentage_min_str:
        try:
            val = float(tenth_percentage_min_str)
            student_profiles_list = student_profiles_list.filter(tenth_percentage__gte=val)
        except ValueError: pass
    if pre_uni_percentage_min_str:
        try:
            val = float(pre_uni_percentage_min_str)
            student_profiles_list = student_profiles_list.filter(pre_university_percentage__gte=val)
        except ValueError: pass
    # --- End: Filtering logic ---

    # --- Create Excel Workbook ---
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Student Placement Data"

    # Define Headers
    headers = [
        "Full Name", "USN", "College Email", "Personal Email", "Phone Number",
        "Department", "Current Year", "Current Semester", 
        "10th %", "12th/Dip Type", "12th/Dip %", "Degree CGPA", "Backlogs", "Batch",
        "Languages Known", "Domains of Interest",
        "LinkedIn", "GitHub", "LeetCode", "HackerRank",
        "Certifications List", "Certificates Drive Link", "Resume"
    ]
    
    # Style for headers (optional)
    header_font = Font(bold=True)
    for col_num, header_title in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        # Adjust column width (optional, can be slow for many columns/rows)
        # column_letter = get_column_letter(col_num)
        # sheet.column_dimensions[column_letter].width = 20 # Example width

    # Populate data rows
    for row_num, profile in enumerate(student_profiles_list, 2): # Start data from row 2
        languages_str = ", ".join([lang.name for lang in profile.languages_known.all()]) or "N/A"
        domains_str = ", ".join([domain.name for domain in profile.domains.all()]) or "N/A"

        row_data = [
            profile.full_name, profile.usn, profile.college_email, profile.personal_email, profile.phone_number,
            profile.get_department_display(), profile.get_current_year_display(), profile.get_current_semester_display(),
            profile.tenth_percentage, profile.get_pre_university_qualification_type_display(), profile.pre_university_percentage,
            profile.cgpa, profile.backlogs, profile.get_batch_display(),
            languages_str, domains_str,
            # For URLs, we'll put the display text and then make them hyperlinks
            "LinkedIn Profile" if profile.linkedin_url else "N/A",
            "GitHub Profile" if profile.github_url else "N/A",
            "LeetCode Profile" if profile.leetcode_url else "N/A",
            "HackerRank Profile" if profile.hackerrank_url else "N/A",
            profile.certifications_list or "N/A",
            "Certificates Drive" if profile.certifications_drive_link else "N/A",
            "View Resume" if profile.resume and profile.resume.url else "N/A",
        ]
        sheet.append(row_data) # Appends to the next available row

        # Add Hyperlinks (current_row_num is sheet.max_row after append)
        current_row_num_in_excel = sheet.max_row
        
        link_map = [ # (Header Name, URL attribute on profile)
            ("LinkedIn", profile.linkedin_url),
            ("GitHub", profile.github_url),
            ("LeetCode", profile.leetcode_url),
            ("HackerRank", profile.hackerrank_url),
            ("Certificates Drive Link", profile.certifications_drive_link),
            ("Resume", profile.resume.url if profile.resume and profile.resume.url else None),
        ]

        for header_name, url_value in link_map:
            if url_value:
                col_idx = headers.index(header_name) + 1 # 1-based column index
                cell = sheet.cell(row=current_row_num_in_excel, column=col_idx)
                try:
                    # For external links, build_absolute_uri is good if the URL is relative
                    # For already absolute URLs (like most profile links), it's fine too
                    absolute_url = request.build_absolute_uri(url_value) if '://' not in str(url_value) and not str(url_value).startswith('/') else str(url_value)
                    cell.hyperlink = absolute_url
                    cell.style = "Hyperlink" # Apply Excel's hyperlink style
                except Exception: # In case build_absolute_uri fails or URL is malformed
                    pass # Leave cell as text

    # --- Prepare HTTP Response ---
    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)

    response = HttpResponse(
        excel_buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"student_placement_data_{timestamp}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response
@login_required(login_url='adminlogin')
def faculty_password_change_view(request):
    if not request.user.is_staff:
        messages.error(request, "This page is for faculty/staff members only.")
        return redirect('home') # Or wherever non-staff should go

    if request.method == 'POST':
        form = PasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important!
            messages.success(request, 'Your password was successfully updated!')
            return redirect('adminportal') # Or back to this page with success, or admin portal
        else:
            messages.error(request, 'Please correct the error(s) below.')
    else:
        form = PasswordChangeForm(user=request.user)
    
    context = {'form': form}
    return render(request, 'faculty_password_change.html', context)
@login_required(login_url='adminlogin')
def faculty_search_student_view(request):
    if not request.user.is_staff:
        messages.error(request, "This page is for faculty/staff members only.")
        return redirect('home')

    student_profile = None
    usn_query = request.GET.get('usn_query', '').strip() # Get USN from GET parameter

    if usn_query:
        try:
            # Search by USN field in StudentProfile or by username in User model
            # Assuming USN is stored in StudentProfile.usn and also User.username
            student_profile = StudentProfile.objects.select_related('user').prefetch_related('domains', 'languages_known').get(usn__iexact=usn_query)
            # Or: student_profile = StudentProfile.objects.get(user__username__iexact=usn_query)
        except StudentProfile.DoesNotExist:
            messages.error(request, f"No student found with USN: {usn_query}")
        except StudentProfile.MultipleObjectsReturned: # Should not happen if USN is unique
            messages.error(request, f"Multiple profiles found for USN: {usn_query}. Please contact admin.")


    context = {
        'student_profile': student_profile,
        'usn_query': usn_query, # To repopulate the search box
    }
    return render(request, 'faculty_search_student.html', context)
def some_view(request):
    raise Exception("This is a test 500 error")
def permission_denied_view(request):
    raise PermissionDenied("You do not have custom permission for this.")
    # ...